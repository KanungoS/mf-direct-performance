# ============================================================
# MUTUAL FUND PERFORMANCE + PORTFOLIO TRACKER ENGINE
# FINAL VERSION (SUPER-STABLE + ALL FEATURES INCLUDED)
# ============================================================

import pandas as pd
import numpy as np
import requests
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_DIR = "data"
CACHE_DIR = f"{DATA_DIR}/cache"
MASTER_CSV = f"{DATA_DIR}/master_list.csv"
DISCONTINUED_CSV = f"{DATA_DIR}/discontinued_schemes.csv"
FINAL_CSV = f"{DATA_DIR}/mf_direct_grid.csv"
FINAL_XLSX = f"{DATA_DIR}/mf_direct_grid.xlsx"
PORTFOLIO_CSV = f"{DATA_DIR}/my_portfolio.csv"

# ============================================================
# UTILITIES
# ============================================================

def safe_mkdir(path):
    if not os.path.exists(path):
        os.makedirs(path)

safe_mkdir(DATA_DIR)
safe_mkdir(CACHE_DIR)

def load_master_list():
    return pd.read_csv(MASTER_CSV)

def save_master_list(df):
    df.to_csv(MASTER_CSV, index=False)

def load_discontinued():
    return pd.read_csv(DISCONTINUED_CSV)

def save_discontinued(df):
    df.to_csv(DISCONTINUED_CSV, index=False)

def fetch_direct_scheme_list():
    url = "https://api.mfapi.in/mf"
    try:
        data = requests.get(url, timeout=20).json()
    except:
        return []
    return [x for x in data if "direct" in x["schemeName"].lower()]

def load_nav_history(code):
    cache_file = f"{CACHE_DIR}/{code}.json"
    url = f"https://api.mfapi.in/mf/{code}"

    if os.path.exists(cache_file):
        js = pd.read_json(cache_file, typ="series")
    else:
        js = requests.get(url, timeout=20).json()
        pd.Series(js).to_json(cache_file)

    if "data" not in js:
        return pd.DataFrame(), js

    df = pd.DataFrame(js["data"])
    df["nav"] = df["nav"].astype(float)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date")
    return df, js

def normalize_category(cat):
    if isinstance(cat, float): return "Other"
    c = cat.lower()
    if "equity" in c: return "Equity"
    if "debt" in c: return "Debt"
    if "hybrid" in c: return "Hybrid"
    if "index" in c: return "Index"
    return cat

def detect_sector_theme(name):
    sectors = {
        "bank": "Banking",
        "pharma": "Pharma",
        "it": "IT",
        "tech": "IT",
        "infra": "Infrastructure",
        "energy": "Energy",
        "consum": "Consumption",
        "fmcg": "Consumption"
    }
    for key, val in sectors.items():
        if key in name.lower():
            return val
    return "Diversified"

# ============================================================
# RETURN CALCULATIONS
# ============================================================

def compute_returns(df):
    if df.empty:
        return {k: None for k in [
            "NAV Latest","NAV 1M","NAV 3M","NAV 1Y","NAV 3Y",
            "Volatility (StdDev 1Y)","Max Drawdown",
            "Rolling Return 1Y","Rolling Return 3Y"
        ]}
    ret = {}
    latest = df.iloc[-1]["nav"]
    ret["NAV Latest"] = latest

    def pct(days):
        if len(df) <= days:
            return None
        try:
            past = df.iloc[-days]["nav"]
            return round((latest - past) * 100 / past, 4)
        except:
            return None

    ret["NAV 1M"] = pct(30)
    ret["NAV 3M"] = pct(90)
    ret["NAV 1Y"] = pct(365)
    ret["NAV 3Y"] = pct(365 * 3)

    # Volatility
    try:
        ret["Volatility (StdDev 1Y)"] = df.tail(365)["nav"].pct_change().std()
    except:
        ret["Volatility (StdDev 1Y)"] = None

    # Max Drawdown
    roll_max = df["nav"].cummax()
    draw = (df["nav"] - roll_max) / roll_max
    ret["Max Drawdown"] = draw.min()

    # Rolling returns
    try:
        ret["Rolling Return 1Y"] = df["nav"].pct_change(365).dropna().mean()
        ret["Rolling Return 3Y"] = df["nav"].pct_change(365 * 3).dropna().mean()
    except:
        ret["Rolling Return 1Y"] = None
        ret["Rolling Return 3Y"] = None

    return ret

# ============================================================
# PEER RANKING
# ============================================================

def compute_category_peer_ranking(df):
    if "NAV 1Y" not in df.columns:
        df["Quartile (1Y)"] = ""
        return df

    df["Category Rank (1Y)"] = df.groupby("Scheme Category")["NAV 1Y"].rank(ascending=False)
    df["Quartile (1Y)"] = df["Category Rank (1Y)"].apply(
        lambda r: (
            "Top Quartile" if r <= 25 else
            "Second Quartile" if r <= 50 else
            "Third Quartile" if r <= 75 else
            "Bottom Quartile"
        )
    )
    return df

def compute_sector_peer_ranking(df):
    df["Sector Rank (1Y)"] = df.groupby("Sector Theme")["NAV 1Y"].rank(ascending=False)
    df["Sector Performance Tag (1Y)"] = df["Sector Rank (1Y)"].apply(
        lambda r: "Sector Leader" if r <= 3 else ("Sector Laggard" if r > 10 else "")
    )
    return df

# ============================================================
# AI FUND SCORE (Aggressive Momentum Model)
# ============================================================

def compute_ai_score(r):
    score = 0

    weights = {
        "NAV 1M": 25,
        "NAV 3M": 25,
        "NAV 1Y": 20,
        "NAV 3Y": 20
    }

    for k, wt in weights.items():
        if pd.notna(r.get(k)):
            score += r[k] * (wt / 100)

    # Volatility penalty
    if pd.notna(r.get("Volatility (StdDev 1Y)")):
        score -= r["Volatility (StdDev 1Y)"] * 80

    # Peer quartile
    if r.get("Quartile (1Y)") == "Top Quartile":
        score += 5

    # Rolling return consistency
    rr = np.nanmean([r.get("Rolling Return 1Y"), r.get("Rolling Return 3Y")])
    if pd.notna(rr):
        score += rr * 10

    return round(score, 2)

# ============================================================
# RISK SCORE
# ============================================================

def compute_risk_score(r):
    vol = r.get("Volatility (StdDev 1Y)")
    dd = r.get("Max Drawdown")

    if vol is None:
        return "Medium"

    if vol < 0.01 and dd > -0.10:
        return "Low"
    if vol < 0.03 and dd > -0.20:
        return "Medium"
    return "High"

# ============================================================
# EXIT LOAD LOGIC
# ============================================================

def compute_exit_load(cat):
    c = cat.lower()
    if "equity" in c:
        return "1% before 1 year"
    if "hybrid" in c:
        return "1% before 90 days"
    if "debt" in c:
        return "0.25% before 30 days"
    if "liquid" in c:
        return "Graded load (7 days)"
    if "elss" in c:
        return "3-year lock-in"
    return ""

# ============================================================
# PORTFOLIO TRACKING
# ============================================================

def load_portfolio():
    if not os.path.exists(PORTFOLIO_CSV):
        df = pd.DataFrame(columns=[
            "Scheme Code","Scheme Name","Units","Purchase NAV","Date of Purchase",
            "Total Purchase Value","Current NAV","Current Value","% Deviation",
            "Will Exit Load Apply","Exit Load %"
        ])
        df.to_csv(PORTFOLIO_CSV, index=False)
        return df
    return pd.read_csv(PORTFOLIO_CSV)

def process_portfolio(port, mf_df):
    out = port.copy()
    out["Total Purchase Value"] = out["Units"] * out["Purchase NAV"]

    out["Current NAV"] = out["Scheme Code"].apply(
        lambda x: mf_df.loc[mf_df["Scheme Code"] == x, "NAV Latest"].iloc[0]
        if x in mf_df["Scheme Code"].values else None
    )

    out["Current Value"] = out["Units"] * out["Current NAV"]
    out["% Deviation"] = ((out["Current NAV"] - out["Purchase NAV"]) / out["Purchase NAV"]) * 100

    # Exit load detection
    def exit_apply(row):
        cat = mf_df.loc[mf_df["Scheme Code"] == row["Scheme Code"], "Scheme Category"].iloc[0]
        rule = compute_exit_load(cat)
        row["Will Exit Load Apply"] = "Yes" if rule else "No"
        row["Exit Load %"] = rule
        return row

    out = out.apply(exit_apply, axis=1)
    return out

# ============================================================
# EXCEL EXPORT
# ============================================================

def save_excel(main_df, portfolio_df):
    wb = Workbook()

    # MAIN SHEET
    ws = wb.active
    ws.title = "MF Grid"
    for r in dataframe_to_rows(main_df, index=False, header=True):
        ws.append(r)
    for col in ws.iter_cols(1, ws.max_column):
        col[0].font = Font(bold=True)

    # PORTFOLIO SHEET
    pws = wb.create_sheet("Portfolio Tracker")
    for r in dataframe_to_rows(portfolio_df, index=False, header=True):
        pws.append(r)
    for col in pws.iter_cols(1, pws.max_column):
        col[0].font = Font(bold=True)

    # TOP 10
    top = wb.create_sheet("Top 10")
    def add_block(title, data):
        top.append([title])
        for r in dataframe_to_rows(data, index=False, header=True):
            top.append(r)
        top.append([""])

    add_block("Top 10 – 1M", main_df.nlargest(10, "NAV 1M"))
    add_block("Top 10 – 3M", main_df.nlargest(10, "NAV 3M"))
    add_block("Top 10 – 1Y", main_df.nlargest(10, "NAV 1Y"))
    add_block("Low Volatility", main_df.nsmallest(10, "Volatility (StdDev 1Y)"))

    wb.save(FINAL_XLSX)

# ============================================================
# MAIN ENGINE
# ============================================================

def main():
    print("Loading master list...")
    master = load_master_list()
    discontinued = load_discontinued()

    print("Fetching current Direct schemes...")
    curr = pd.DataFrame(fetch_direct_scheme_list())

    prev = set(master["Scheme Code"])
    now = set(curr["schemeCode"])

    removed = prev - now
    added = now - prev

    # mark discontinued
    for c in removed:
        old = master.loc[master["Scheme Code"] == c, "Scheme Name"].iloc[0]
        discontinued = pd.concat([
            discontinued,
            pd.DataFrame([{
                "Scheme Code": c,
                "Scheme Name": old,
                "Date Discontinued": datetime.today().strftime("%Y-%m-%d")
            }])
        ])

    # add new schemes
    for c in added:
        nm = curr[curr["schemeCode"] == c]["schemeName"].iloc[0]
        master = pd.concat([
            master,
            pd.DataFrame([{
                "Scheme Code": c,
                "Scheme Name": nm,
                "Scheme Category": "",
                "Scheme Status": "Active"
            }])
        ])

    rows = []

    print("\nProcessing NAV + analytics:\n")
    for i, r in master.iterrows():
        code = r["Scheme Code"]
        name = r["Scheme Name"]

        print(f" → {i+1}/{len(master)}  {name}")

        df, js = load_nav_history(code)
        meta = js.get("meta", {})

        cat = normalize_category(meta.get("category", ""))
        amc = meta.get("fund_house", "")
        sector = detect_sector_theme(name)

        ret = compute_returns(df)

        rows.append({
            "AMC": amc,
            "Scheme Code": code,
            "Scheme Name": name,
            "Scheme Category": cat,
            "Sector Theme": sector,
            **ret
        })

    mf_df = pd.DataFrame(rows)

    # Rankings
    mf_df = compute_category_peer_ranking(mf_df)
    mf_df = compute_sector_peer_ranking(mf_df)

    # AI score
    mf_df["AI Fund Score"] = mf_df.apply(compute_ai_score, axis=1)

    # Risk score
    mf_df["Risk Score"] = mf_df.apply(compute_risk_score, axis=1)

    # Exit load
    mf_df["Exit Load"] = mf_df["Scheme Category"].apply(compute_exit_load)

    # Save main CSV
    mf_df.to_csv(FINAL_CSV, index=False)

    # Portfolio processing
    portfolio = load_portfolio()
    portfolio = process_portfolio(portfolio, mf_df)

    # Save Excel
    save_excel(mf_df, portfolio)

    # Save master files
    save_master_list(master)
    save_discontinued(discontinued)

    print("\nDONE ✓ All outputs generated.")


if __name__ == "__main__":
    main()
